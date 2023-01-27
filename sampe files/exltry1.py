from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root=Tk()
root.title("excel load")
root.geometry("500x800")

wb=Workbook()

wb=load_workbook("samplexl.xlsx")

ws=wb.active

column_a=ws['A']
column_b=ws['B']
def get_a():
     list=''
     for cell in column_a:
          list=f'{list+str(cell.value)}\n'
          
def get_b():
     list=''
     for cell in column_b:
          list=f'{list+str(cell.value)}\n'

root.mainloop()
