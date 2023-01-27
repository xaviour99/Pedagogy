from tkinter import *
from tkinter import messagebox as ms
from PIL import ImageTk, Image 
import tkinter as tk
from PIL import Image
from tkinter import Message ,Text
import cv2
import os
import shutil
import csv
import numpy as np
from PIL import Image, ImageTk
import pandas as pd
import datetime
import time
import tkinter.ttk as ttk
import tkinter.font as font
import sqlite3
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from tkinter import filedialog
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pickle

global secret_key
secret_key="sreedhar"


with sqlite3.connect('Datas/Account_database.db') as db:
    c = db.cursor()

c.execute('CREATE TABLE IF NOT EXISTS user (username TEXT NOT NULL PRIMARY KEY,password TEXT NOT NULL,category TEXT  NOT NULL)')
db.commit()
db.close()


class main:
    def __init__(self,master):
    	# Window 
        self.master = master
        # Some Usefull variables
        self.username = StringVar()
        self.password = StringVar()
        self.n_username = StringVar()
        self.n_password = StringVar()
        self.n_category = StringVar()
        self.n_secret_key = StringVar()
        #Create Widgets
        self.widgets()

    #Login Function
    def login(self):
    	#Establish Connection
        with sqlite3.connect('Datas/Account_database.db') as db:
            c = db.cursor()

        #Find user If there is any take proper action
        find_user = ('SELECT category FROM user WHERE username = ? and password = ?')
        c.execute(find_user,[(self.username.get()),(self.password.get())])
        result = c.fetchall()
        if result:
            self.logf.pack_forget()
            self.head['text'] = self.username.get() + '\n Logged In'
            self.head['pady'] = 150
            def staff_window():

                stf1_window=Toplevel(root)
                stf1_window.title("Staff Home Page")

                             
                stf1_window.geometry('1280x700')
                stf1_window.iconbitmap('Datas/src_images/face5.ico')
                stf1_window.state('zoomed')  
                #stf1_window.resizable(False, False)
                stf1_window.configure(background='gray')
                stf1_window.grid_rowconfigure(0, weight=1)
                stf1_window.grid_columnconfigure(0, weight=1)
                load = Image.open('Datas/src_images/face_reco.jpg')
                render = ImageTk.PhotoImage(load)
                img = Label(stf1_window, image=render)
                img.place(x=0, y=0, relwidth=1, relheight=1)

                video_upload_photo = ImageTk.PhotoImage(file = "Datas/src_images/upload_btn_logo1.png")

                pdf_upload_photo = ImageTk.PhotoImage(file = "Datas/src_images/upload_pdf_logo.png")

                stf_study_upload_photo = ImageTk.PhotoImage(file = "Datas/src_images/upload_pdf_logo_stm.png")







                message = Label(stf1_window, text="Welcome Staff : "+self.username.get() ,bg="lightblue"  ,fg="black"  ,width=45  ,height=1,font=('calibri', 40, 'bold')) 
                message.pack(side=TOP)
                #message.place(x=465, y=0)

                def test():
                    window=Toplevel(stf1_window)
                    window.title("Face_Recogniser")

                    dialog_title = 'QUIT'
                    dialog_text = 'Are you sure?'             
                    window.geometry('1280x650')
                    window.iconbitmap('Datas/src_images/face5.ico')
                    window.resizable(False, False)
                    window.configure(background='gray')
                    window.grid_rowconfigure(0, weight=1)
                    window.grid_columnconfigure(0, weight=1)
                    load = Image.open('Datas/src_images/face6.jpg')
                    render = ImageTk.PhotoImage(load)
                    img = Label(window, image=render)
                    img.place(x=0, y=0, relwidth=1, relheight=1)

                    message = Label(window, text=" Student Face Entry" ,bg="lightblue"  ,fg="black"  ,width=18  ,height=1,font=('calibri', 30, 'bold')) 
                    message.place(x=465, y=0)
                    lbl = Label(window, text="Enter ID",width=13  ,height=2  ,fg="black"  ,bg="lightblue" ,font=('calibri', 15, ' bold ') ) 
                    lbl.place(x=370, y=150)

                    txt = Entry(window,width=20  ,bg="lightblue" ,fg="black",font=('calibri', 15, ' bold '))
                    txt.place(x=600, y=165)

                    lbl2 = Label(window, text="Enter Name",width=13  ,fg="black"  ,bg="lightblue"    ,height=2 ,font=('calibri', 15, ' bold ')) 
                    lbl2.place(x=370, y=250)

                    txt2 = Entry(window,width=20  ,bg="lightblue"  ,fg="black",font=('calibri', 15, ' bold '))
                    txt2.place(x=600, y=265)

                    lbl3 = Label(window, text="Notification : ",width=13  ,fg="black"  ,bg="lightblue"  ,height=2 ,font=('calibri', 15, ' bold ')) 
                    lbl3.place(x=370, y=350)

                    message = Label(window, text="" ,bg="lightblue"  ,fg="black"  ,width=30  ,height=2, activebackground = "lightblue" ,font=('calibri', 15, ' bold ')) 
                    message.place(x=600, y=355)

                    # lbl3 = Label(window, text="Attendance : ",width=20  ,fg="red"  ,bg="lightblue"  ,height=2 ,font=('calibri', 15, ' bold  underline')) 
                    # lbl3.place(x=400, y=650)


                    # message2 = Label(window, text="" ,fg="red"   ,bg="lightblue",activeforeground = "green",width=30  ,height=2  ,font=('calibri', 15, ' bold ')) 
                    # message2.place(x=700, y=650)
                      
                    def clear():
                        txt.delete(0, 'end')    
                        res = ""
                        message.configure(text= res)

                    def clear2():
                        txt2.delete(0, 'end')    
                        res = ""
                        message.configure(text= res)    
                        
                    def is_number(s):
                        try:
                            float(s)
                            return True
                        except ValueError:
                            pass
                     
                        try:
                            import unicodedata
                            unicodedata.numeric(s)
                            return True
                        except (TypeError, ValueError):
                            pass
                     
                        return False
                
                    def TakeImages():        
                        Id=(txt.get())
                        name=(txt2.get())
                        if(is_number(Id) and name.isalpha()):
                            cam = cv2.VideoCapture(0)
                            recognizer = cv2.face.LBPHFaceRecognizer_create()

                            #recognizer = cv2.face.LBPHFaceRecognizer_create()
                            harcascadePath ="Datas/haarcascade_frontalface_default.xml"
                            detector =cv2.CascadeClassifier(harcascadePath)
                            sampleNum=0
                            while(True):
                                ret, img = cam.read()
                                gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                                faces=detector.detectMultiScale(gray,1.3,5)
                                for (x,y,w,h) in faces:
                                    cv2.rectangle(img,(x,y),(x+w,y+h),(255,0,0),2)        
                                    sampleNum=sampleNum+1
                                    cv2.imwrite("Atten_datas/TrainingImage/ "+name +"."+Id +'.'+ str(sampleNum) + ".jpg", gray[y:y+h,x:x+w])
                                   
                                    
                                    cv2.imshow('frame',img)
                                
                                if cv2.waitKey(100) & 0xFF == ord('q'):
                                    break
                                
                                elif sampleNum>60:
                                    break
                            cam.release()
                            cv2.destroyAllWindows() 
                            res = "Images Saved for ID : " + Id +" Name : "+ name
                            row = [Id , name]
                            with open('Atten_datas/AuthorisedDetails/AuthorisedDetails.csv','a+') as csvFile:
                                writer = csv.writer(csvFile)
                                writer.writerow(row)
                            csvFile.close()
                            message.configure(text= res)
                        else:
                            if(is_number(Id)):
                                res = "Enter Alphabetical Name"
                                message.configure(text= res)
                            if(name.isalpha()):
                                res = "Enter Numeric Id"
                                message.configure(text= res)
                        
                    def TrainImages():
                        recognizer = cv2.face.LBPHFaceRecognizer_create()                        
                        
                        harcascadePath ="Datas/haarcascade_frontalface_default.xml" 
                        detector =cv2.CascadeClassifier(harcascadePath)
                        faces,Id = getImagesAndLabels("Atten_datas/TrainingImage")
                        recognizer.train(faces, np.array(Id))
                        recognizer.save("Atten_datas/TrainingImageLabel/Trainner.yml")
                        res = "Image Trained"
                        message.configure(text= res)

                    def getImagesAndLabels(path):
                        
                        imagePaths=[os.path.join(path,f)for f in os.listdir(path)] 
                        
                        
                        
                        faces=[]
                        
                        Ids=[]
                        
                        for imagePath in imagePaths:
                            
                            pilImage=Image.open(imagePath).convert('L')
                            
                            imageNp=np.array(pilImage,'uint8')
                            
                            Id=int(os.path.split(imagePath)[-1].split(".")[1])
                            
                            faces.append(imageNp)
                            Ids.append(Id)        
                        return faces,Ids

                    def TrackImages():
                        recognizer = cv2.face.LBPHFaceRecognizer_create()
                        recognizer.read("Atten_datas/TrainingImageLabel/Trainner.yml")
                        harcascadePath="Datas/haarcascade_frontalface_default.xml"
                        faceCascade = cv2.CascadeClassifier(harcascadePath);    
                        df=pd.read_csv('Atten_datas/AuthorisedDetails/AuthorisedDetails.csv')
                        cam = cv2.VideoCapture(0)
                        font = cv2.FONT_HERSHEY_SIMPLEX        
                        col_names =  ['Id','Name','Date','Time']
                        attendance = pd.DataFrame(columns = col_names)    
                        while True:
                            ret, im =cam.read()
                            gray=cv2.cvtColor(im,cv2.COLOR_BGR2GRAY)
                            faces=faceCascade.detectMultiScale(gray, 1.2,5)    
                            for(x,y,w,h) in faces:
                                cv2.rectangle(im,(x,y),(x+w,y+h),(225,0,0),2)
                                Id, conf = recognizer.predict(gray[y:y+h,x:x+w])                                   
                                if(conf < 50):
                                    ts = time.time()      
                                    date = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
                                    calibritamp = datetime.datetime.fromtimestamp(ts).strftime('%H:%M:%S')
                                    aa=df.loc[df['Id'] == Id]['Name'].values
                                    tt=str(Id)+"-"+aa[0]
                                    attendance.loc[len(attendance)] = [Id,aa[0],date,calibritamp]
                                    
                                else:
                                    Id='Unknown'                
                                    tt=str(Id)  
                                if(conf > 75):
                                    Id='Unknown'                
                                    tt=str(Id)                
                                    
                                cv2.putText(im,str(tt),(x,y+h), font, 1,(255,255,255),2)        
                            attendance=attendance.drop_duplicates(subset=['Id'],keep='first')    
                            cv2.imshow('image Detected',im) 
                            if (cv2.waitKey(1)==ord('q')):
                                break
                        ts = time.time()      
                        date = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
                        calibritamp = datetime.datetime.fromtimestamp(ts).strftime('%H:%M:%S')
                        Hour,Minute,Second=calibritamp.split(":")
                        fileName="Atten_datas/Attendance_sheet/"+"Attendance_"+date+"_"+Hour+"-"+Minute+"-"+Second+".csv"
                        attendance.to_csv(fileName,index=False)
                        cam.release()
                        cv2.destroyAllWindows()
                        print(attendance)
                        #res=attendance
                        #message2.configure(text= res)



                    clearButton = Button(window, text="Clear", command=clear  ,fg="black"  ,bg="lightblue"  ,width=9  ,height=0 ,activebackground = "Red" ,font=('calibri', 15, ' bold '))
                    clearButton.place(x=870, y=155)
                    clearButton2 = Button(window, text="Clear", command=clear2  ,fg="black"  ,bg="lightblue"  ,width=9  ,height=0, activebackground = "Red" ,font=('calibri', 15, ' bold '))
                    clearButton2.place(x=870, y=255)    
                    takeImg = Button(window, text="Take Images", command=TakeImages  ,fg="black"  ,bg="lightblue"  ,width=12  ,height=1, activebackground = "Red" ,font=('calibri', 15, ' bold '))
                    takeImg.place(x=300, y=585)
                    trainImg = Button(window, text="Train Images", command=TrainImages  ,fg="black"  ,bg="lightblue"  ,width=12  ,height=1, activebackground = "Red" ,font=('calibri', 15, ' bold '))
                    trainImg.place(x=500, y=585)
                    trackImg = Button(window, text="Track Images", command=TrackImages  ,fg="black"  ,bg="lightblue"  ,width=12  ,height=1, activebackground = "Red" ,font=('calibri', 15, ' bold '))
                    trackImg.place(x=700, y=585)
                    quitWindow = Button(window, text="Quit", command=window.destroy  ,fg="white"  ,bg="#e82902"  ,width=12  ,height=1, activebackground = "Red" ,font=('calibri', 15, ' bold '))
                    quitWindow.place(x=900, y=585)
                    
                    window.mainloop()



                def stf_show_att():
                    stf_atten_report_window = Toplevel(stf1_window)
                    stf_atten_report_window.title("Students Attendance_sheet")

                    stf_atten_report_window.geometry("600x450")
                    stf_atten_report_window.iconbitmap('Datas/src_images/face5.ico')
                    stf_atten_report_window.resizable(False, False)
                    stf_atten_report_window.configure(background='gray')
                    Download_logo_photo = PhotoImage(file = r"Datas/src_images/download-logo-12.png")

                    scroll_bar = Scrollbar(stf_atten_report_window)
                    scroll_bar.pack( side = RIGHT,fill = Y )
                    mylist = Listbox(stf_atten_report_window,font = ('calibri', 22, ' bold '),background='gray',yscrollcommand = scroll_bar.set )



                    stud_atten_report_files=os.listdir('Atten_datas/Attendance_sheet')
                    for linenum,line in enumerate(stud_atten_report_files):
                        mylist.insert(END,str(linenum+1)+'. '+str(line))
                        mylist.insert(END,'\n')
                    mylist.pack( side = TOP, fill = BOTH ,expand=True )
                    scroll_bar.config( command = mylist.yview )
                    frame = Frame(stf_atten_report_window)
                    frame.pack(fill=X,side=BOTTOM)
                    lbt1=Label(frame,text="Click here-->",font = ('calibri', 22, ' bold '))
                    lbt1.pack(fill = X,side=LEFT)

                        
                    def predicting_atten(filename):
                         atten_data=pd.read_csv(filename)
                         atten_data.dropna(inplace=True)
                         id_data=list(atten_data['Id'])
                         name_data=list(atten_data['Name'])
                         date_data=list(atten_data['Date'])
                         time_data=list(atten_data['Time'])                         
                         
                         predicting_atten_window=Toplevel(stf_atten_report_window)
                         predicting_atten_window.geometry("1200x600")
                         predicting_atten_window.iconbitmap('Datas/src_images/face5.ico')
                         predicting_atten_window.resizable(False, False)
                         predicting_atten_window.configure(background='gray')
                         frame = Frame(predicting_atten_window)
                         frame.pack(fill=X,side=TOP)
                         lbt1=Label(frame,text="Attendance Report",font=("Imprint MT Shadow",35))
                         lbt1.pack(fill = X, expand = True,side=TOP)
                         frame1 = Frame(predicting_atten_window)
                         frame1.pack(fill=X,side=TOP)
                         lbl1=Label(frame1,text="Id              ",font=('calibri', 30, ' bold '))
                         lbl1.pack(side=LEFT)
                         lbl2=Label(frame1,text="Time                       ",font=('calibri', 30, ' bold '))
                         lbl2.pack(side=LEFT)
                         lbl3=Label(frame1,text="Date                       ",font=('calibri', 30, ' bold '))
                         lbl3.pack(side=LEFT)
                         lbl4=Label(frame1,text="Name                       ",font=('calibri', 30, ' bold '))
                         lbl4.pack(side=LEFT)
                         
                       
                         
                         
                         
                         scroll_bar = Scrollbar(predicting_atten_window)     

                         mylist = Listbox(predicting_atten_window,font = ('calibri', 26, ' bold '),background='gray',yscrollcommand = scroll_bar.set)

                         atten_list=[]
                         atten_r=[]
                         total_stud_details=pd.read_csv(r'Atten_datas/AuthorisedDetails/AuthorisedDetails.csv')
                         total_stud_names=list(total_stud_details['Name'])
                         
                         for iname in total_stud_names:
                              if iname in name_data:
                                   id_v=atten_data.loc[atten_data['Name'] == iname]['Id'].values
                                   time_val=atten_data.loc[atten_data['Name'] == iname]['Time'].values
                                   date_val=atten_data.loc[atten_data['Name'] == iname]['Date'].values
                                   atten_r.append("Present")               
                                   
                                   mylist.insert(END,str(id_v[0])+'              '+str(time_val)+'                      '+str(date_val)+'              '+str(iname))
                                   mylist.insert(END,'\n')
                              elif iname:
                                   id_v2=total_stud_details.loc[total_stud_details['Name'] == iname]['Id'].values
                                   if id_v2.size>0:
                                    mylist.insert(END,str(int(id_v2[0]))+'              '+str('Unknown')+'                      '+str('Unknown')+'              '+str(iname))
                                    mylist.insert(END,'\n')
                                    atten_r.append("Absent")

                         atten_r_d={"Atten":atten_r}                      
                                   
  
                         
                         def pred_atten_report_graph():
                              
                              num = pd.value_counts(atten_r_d['Atten'],sort=True).sort_index()
                              num.plot(kind='bar')
                              plt.title("Student Attendance Report ")
                              plt.xlabel("Attendance")
                              plt.ylabel("no of students")
                              plt.rcParams['figure.figsize'] = [10,4]
                              plt.show()

                         frame2 = Frame(predicting_atten_window)
                         frame2.pack(fill=X,side=TOP)
                         
                         
                         blbl1=Label(frame2,text="To View graph the total Attendance report",font=('calibri', 20, ' bold '),fg="black"  ,bg="lightblue", activebackground = "#002266",activeforeground="#80b3ff" )
                         blbl1.pack(fill = X, expand = True,side=LEFT,padx=10)
                         blbl2=Button(frame2,text="show graph",command=pred_atten_report_graph,font=('calibri', 20, ' bold '),fg="black"  ,bg="lightblue", activebackground = "#002266",activeforeground="#80b3ff")
                         blbl2.pack(fill = X, expand = True,side=LEFT,padx=10)
                         

                         scroll_bar.pack( side = RIGHT,fill = Y )
                         mylist.pack( side = TOP, fill = BOTH ,expand=True )
                         scroll_bar.config( command = mylist.yview )
                         
                         predicting_atten_window.mainloop()

                    def openfilename_atten_mk():
                         filename='Atten_datas/Attendance_sheet/'+clicked.get()

                         if filename.endswith('.csv'):
                              csv_stu_data=pd.read_csv(filename)
                              csv_stu_data_col=list(csv_stu_data.columns)

                              if 'Id' in csv_stu_data_col and 'Name' in csv_stu_data_col and 'Date' in csv_stu_data_col and 'Time':

                                   predicting_atten(filename)
                              else:
                                   ms.showerror('Error!','Selected file does not have required colums \n Id, Name, Date,Time.',parent=stf_atten_report_window)  
                         else:
                              ms.showerror('Error!','Selected file is not csv file \nMake sure it is csv file.',parent=stf_atten_report_window)


                    stasgnup_btn1=Button(frame,image=Download_logo_photo,command=openfilename_atten_mk)
                    stasgnup_btn1.pack(side=LEFT)
                    clicked = StringVar()
                    clicked.set( stud_atten_report_files[-1] )
                    drop_atten = OptionMenu( frame , clicked , *stud_atten_report_files )
                    drop_atten.config( font=('calibri', 20, ' bold '))
                    drop_atten.pack(fill = X,side=LEFT)
                    stf_atten_report_window.mainloop()
                def stf_record():
                    stf_record_window=Toplevel(stf1_window)
                    stf_record_window.geometry('600x400')
                    stf_record_window.title('Video Upload')
                    stf_record_window.iconbitmap('Datas/src_images/face5.ico')
                    stf_record_window.resizable(False, False)
                    stf_record_window.configure(background='gray')
                    
                    scroll_bar = Scrollbar(stf_record_window)

                    scroll_bar.pack( side = RIGHT,fill = Y )

                    mylist = Listbox(stf_record_window,font = ('calibri', 22, ' bold '),background='gray',yscrollcommand = scroll_bar.set )

                    video_files=os.listdir('Datas/videos')

                    for linenum,line in enumerate(video_files):
                        mylist.insert(END,str(linenum+1)+'. '+line)
                        mylist.insert(END,'\n')


                    def openfile_record():
                         filename = filedialog.askopenfilename(parent=stf_record_window,title ="Select Video")
                         
                         if filename:
                              upvideofile=filename.split('/')
                              if upvideofile[-1] in video_files:
                                   ms.showerror('Error!','file Already exists',parent=stf_record_window)
                                   
                              elif filename.endswith('.mp4') or filename.endswith('.mkv'):
                                   shutil.copy(filename, 'Datas/videos')
                                   cr_video_files=os.listdir('Datas/videos')                                   
                                   ms.showinfo('Success!',upvideofile[-1]+' Upload Successfully!',parent=stf_record_window)
                                   mylist.insert(END,str(len(cr_video_files))+'. '+str(upvideofile[-1]))
                                   mylist.insert(END,'\n')
                              else:
                                   ms.showerror('Error!','Selected file is not video file \nMake sure it is mp4 or mkv file.',parent=stf_record_window)



                    mylist.pack( side = TOP, fill = BOTH ,expand=True)
                    scroll_bar.config( command = mylist.yview )

                    rc_btn1=Button(stf_record_window,image=video_upload_photo,command=openfile_record)
                    rc_btn1.place(x=480,y=300)


                    stf_record_window.mainloop()
                    

                def stf_class():
                    staff_zoom_cls_shed_window=Toplevel(stf1_window)
                    staff_zoom_cls_shed_window.title("Staff class assign Page")

                                                 
                    staff_zoom_cls_shed_window.geometry('1000x500')
                    staff_zoom_cls_shed_window.iconbitmap('Datas/src_images/face5.ico')
                    staff_zoom_cls_shed_window.resizable(False, False)
                    staff_zoom_cls_shed_window.configure(background='gray')
                    staff_zoom_cls_shed_window.grid_rowconfigure(0, weight=1)
                    staff_zoom_cls_shed_window.grid_columnconfigure(0, weight=1)

                    load = Image.open('Datas/src_images/face6.jpg')
                    render = ImageTk.PhotoImage(load)

                    img = Label(staff_zoom_cls_shed_window, image=render)
                    img.place(x=0, y=0, relwidth=1, relheight=1)
                    linkv=StringVar()
                    sub_clicked = StringVar()
                    dtm = IntVar()

                    htm=IntVar()

                    mtm=IntVar()

                    def sheduling_meet_manual():
                        def assigning_the_link(xlfilename):

                            f1=pd.read_excel((xlfilename),engine='openpyxl')

                            current_date_and_time = datetime.datetime.now()

                            hours_added = datetime.timedelta(days = dtm.get(),hours=htm.get(),minutes=mtm.get())
                            future_date_and_time = current_date_and_time + hours_added

                            df = pd.DataFrame({'Time(in format: dd-mm-yyyy hh:mm AM/PM)':[future_date_and_time.strftime("%d-%m-%Y %I:%M %p")],'link':[linkv.get()]})
                            writer = pd.ExcelWriter(xlfilename, engine='openpyxl')
                            # try to open an existing workbook
                            writer.book = load_workbook(xlfilename)
                            # copy existing sheets
                            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                            # read existing file
                            reader = pd.read_excel(open(xlfilename, 'rb'))
                            # write out the new sheet
                            df.to_excel(writer,index=False,header=False,startrow=len(reader)+1)

                            writer.close()
                            ms.showinfo('Success!',' Upload Successfully!',parent=staff_zoom_cls_shed_window)
                            ent1.delete(first=0,last=END)

                        subjectclick=sub_clicked.get()

                        subjectclickdict={"Java":"Datas/zoom_meeting_items/zoom_meet_xlfiles/Java_meet.xlsx" ,
                        "Web Technology":"Datas/zoom_meeting_items/zoom_meet_xlfiles/Web_Technology_meet.xlsx",
                        "Software testing":"Datas/zoom_meeting_items/zoom_meet_xlfiles/Software testing_meet.xlsx",
                        "Mobile app development":"Datas/zoom_meeting_items/zoom_meet_xlfiles/Mobile_app_development_meet.xlsx",
                        "Fundamentals of datascience":"Datas/zoom_meeting_items/zoom_meet_xlfiles/Fundamentals_of_datascience_meet.xlsx",
                        "Data Structures":"Datas/zoom_meeting_items/zoom_meet_xlfiles/Data_Structures_meet.xlsx"}
                        
                        subjexl=subjectclickdict[subjectclick]
                        
                        assigning_the_link(subjexl)




                    staff_zoom_cls_shed_label = Label(staff_zoom_cls_shed_window, text="Shedule the class timing" ,bg="lightblue"  ,fg="black"  ,width=45  ,height=1,font=('calibri', 40, 'bold')) 
                    staff_zoom_cls_shed_label.pack(side=TOP)
                    staff_zoom_cls_shed_label5 = Label(staff_zoom_cls_shed_window, text="Link :" ,bg="lightblue"  ,fg="black"  ,width=5  ,height=1,font=('calibri', 22, 'bold')) 
                    staff_zoom_cls_shed_label5.place(x=100, y=150)

                    staff_zoom_cls_shed_label6 = Label(staff_zoom_cls_shed_window, text="Link Activation Time Period" ,bg="lightblue"  ,fg="black"  ,width=25  ,height=1,font=('calibri', 22, 'bold')) 
                    staff_zoom_cls_shed_label6.place(x=350, y=250)

                    staff_zoom_cls_shed_label2 = Label(staff_zoom_cls_shed_window, text="   Day   " ,bg="lightblue"  ,fg="black"   ,height=1,font=('calibri', 22, 'bold')) 
                    staff_zoom_cls_shed_label2.place(x=200, y=300)
                    staff_zoom_cls_shed_label3 = Label(staff_zoom_cls_shed_window, text="  Hour   " ,bg="lightblue"  ,fg="black"   ,height=1,font=('calibri', 22, 'bold')) 
                    staff_zoom_cls_shed_label3.place(x=450, y=300)
                    staff_zoom_cls_shed_label4 = Label(staff_zoom_cls_shed_window, text=" Minutes " ,bg="lightblue"  ,fg="black"   ,height=1,font=('calibri', 22, 'bold')) 
                    staff_zoom_cls_shed_label4.place(x=700, y=300)


                    assgn_sub_shed_names=['Java',"Web Technology","Software testing","Mobile app development","Fundamentals of datascience","Data Structures"]

                    ent1=Entry(staff_zoom_cls_shed_window,textvariable=linkv ,bg="lightblue"  ,fg="black"  ,width=45  ,font=('calibri', 22, 'bold'))
                    ent1.place(x=200, y=150)

                    ent2 = Scale( staff_zoom_cls_shed_window,font=('calibari',22),bg='lightblue', variable = dtm,from_ = 0, to = 30,
                                    resolution = 1,orient = HORIZONTAL,sliderlength=30,troughcolor='lightblue')
                    ent2.place(x=200, y=350,width=150)

                    ent3 = Scale( staff_zoom_cls_shed_window,font=('calibari',22),bg='lightblue', variable = htm,from_ = 1, to = 24,
                                    resolution = 1,orient = HORIZONTAL,sliderlength=30,troughcolor='lightblue')
                    ent3.place(x=450, y=350,width=150)

                    ent4 = Scale( staff_zoom_cls_shed_window,font=('calibari',22),bg='lightblue', variable = mtm,from_ = 0, to = 24,
                                    resolution = 1,orient = HORIZONTAL,sliderlength=30,troughcolor='lightblue')
                    ent4.place(x=700, y=350,width=150)

                    frame = Frame(staff_zoom_cls_shed_window,bg='grey')
                    frame.pack(fill=X,side=BOTTOM)

                    stasgnup_btn1=Button(frame,command=sheduling_meet_manual,text="Start Session",font=('calibri', 20, ' bold '),fg="black"  ,bg="lightblue"  ,width=15  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" )
                    stasgnup_btn1.pack(side=RIGHT,padx=15)

                    lbt1=Label(frame,text="<--Select Subject",font = ('calibri', 22, ' bold '),bg='grey')
                    lbt1.pack(fill = X,side=RIGHT)


                    sub_clicked.set( assgn_sub_shed_names[-1] )
                    drop_assgn = OptionMenu( frame , sub_clicked , *assgn_sub_shed_names )
                    drop_assgn.config( font=('calibri', 20, ' bold '))
                    drop_assgn.pack(fill = X,side=RIGHT)
                 
                    staff_zoom_cls_shed_window.mainloop()

                def stf_assignment():
                    stf_assignments_up_window = Toplevel(stf1_window)
                    stf_assignments_up_window.title("Assignments Upload")

                    stf_assignments_up_window.geometry("600x400")
                    stf_assignments_up_window.iconbitmap('Datas/src_images/face5.ico')
                    stf_assignments_up_window.resizable(False, False)
                    stf_assignments_up_window.configure(background='gray')
                     
                    scroll_bar = Scrollbar(stf_assignments_up_window)

                    scroll_bar.pack( side = RIGHT,fill = Y )

                    mylist = Listbox(stf_assignments_up_window,font = ('calibri', 22, ' bold '),background='gray',yscrollcommand = scroll_bar.set )

                    Assignments_pdf_files=os.listdir('Datas/Assignments_Upload')

                    for linenum,line in enumerate(Assignments_pdf_files):
                        mylist.insert(END,str(linenum+1)+'. '+str(line))
                        mylist.insert(END,'\n')


                    def openfile_up_Assignments_pdf():
                         filename = filedialog.askopenfilename(parent=stf_assignments_up_window,title ="Select pdf")
                         
                         if filename:
                              uppdffile=filename.split('/')
                              if uppdffile[-1] in Assignments_pdf_files:
                                   ms.showerror('Error!','file Already exists',parent=stf_assignments_up_window)
                                   
                              elif filename.endswith('.pdf'):
                                   shutil.copy(filename, 'Datas/Assignments_Upload')
                                   cr_Assignments_pdf_files=os.listdir('Datas/Assignments_Upload')
                                   
                                   
                                   ms.showinfo('Success!',uppdffile[-1]+' Upload Successfully!',parent=stf_assignments_up_window)
                                   mylist.insert(END,str(len(cr_Assignments_pdf_files))+'. '+str(uppdffile[-1]))
                                   mylist.insert(END,'\n')
                              else:
                                   ms.showerror('Error!','Selected file is not pdf file \nMake sure it is pdf.',parent=stf_assignments_up_window)



                    mylist.pack( side = TOP, fill = BOTH ,expand=True )
                    scroll_bar.config( command = mylist.yview )

                    asn_btn1=Button(stf_assignments_up_window,image=pdf_upload_photo,command=openfile_up_Assignments_pdf)
                    asn_btn1.place(x=480,y=300)


                    stf_assignments_up_window.mainloop()
                def stf_study_material():
                    stf_study_materials_window = Toplevel(stf1_window)
                    stf_study_materials_window.title("Upload study materials")

                    stf_study_materials_window.geometry("600x400")
                    stf_study_materials_window.iconbitmap('Datas/src_images/face5.ico')
                    stf_study_materials_window.resizable(False, False)
                    stf_study_materials_window.configure(background='gray')
                    
                     
                    scroll_bar = Scrollbar(stf_study_materials_window)

                    scroll_bar.pack( side = RIGHT,fill = Y )

                    mylist = Listbox(stf_study_materials_window,font = ('calibri', 22, ' bold '),background='gray',yscrollcommand = scroll_bar.set )

                    pdf_files=os.listdir('Datas/study_materials')

                    for linenum,line in enumerate(pdf_files):
                        mylist.insert(END,str(linenum+1)+'. '+str(line))
                        mylist.insert(END,'\n')


                    def openfile_up_pdf():
                         filename = filedialog.askopenfilename(parent=stf_study_materials_window,title ="Select pdf")
                         
                         if filename:
                              uppdffile=filename.split('/')
                              if uppdffile[-1] in pdf_files:
                                   ms.showerror('Error!','file Already exists',parent=stf_study_materials_window)
                                   
                              elif filename.endswith('.pdf'):
                                   shutil.copy(filename, 'Datas/study_materials')
                                   cr_pdf_files=os.listdir('Datas/study_materials')
                                   
                                   
                                   ms.showinfo('Success!',uppdffile[-1]+' Upload Successfully!',parent=stf_study_materials_window)
                                   mylist.insert(END,str(len(cr_pdf_files))+'. '+str(uppdffile[-1]))
                                   mylist.insert(END,'\n')
                              else:
                                   ms.showerror('Error!','Selected file is not pdf file \nMake sure it is pdf.',parent=stf_study_materials_window)



                    mylist.pack( side = TOP, fill = BOTH ,expand=True )
                    scroll_bar.config( command = mylist.yview )

                    stm_btn1=Button(stf_study_materials_window,image=stf_study_upload_photo,command=openfile_up_pdf)
                    stm_btn1.place(x=480,y=300)


                    stf_study_materials_window.mainloop()

                def stf_leaderboard():
                    stf_leaderboard_window = Toplevel(stf1_window)
                    stf_leaderboard_window.title("Result prediction and Analysis")

                    stf_leaderboard_window.geometry("600x400")
                    stf_leaderboard_window.iconbitmap('Datas/src_images/face5.ico')
                    stf_leaderboard_window.resizable(False, False)
                    stf_leaderboard_window.configure(background='gray')
                    import_csv_photo = PhotoImage(file = r"Datas/src_images/import-csv.png") 
                    scroll_bar = Scrollbar(stf_leaderboard_window)

                    scroll_bar.pack( side = RIGHT,fill = Y )

                    mylist = Listbox(stf_leaderboard_window,font = ('calibri', 22, ' bold '),background='gray',yscrollcommand = scroll_bar.set)

                    mark_files=os.listdir('Datas/mark_data')

                    for linenum,line in enumerate(mark_files):
                        mylist.insert(END,str(linenum+1)+'. '+str(line))
                        mylist.insert(END,'\n')

                        
                    def predicting_mark(filename):
                         mark_data=pd.read_csv(filename)
                         mark_data.dropna(inplace=True)
                         regno_data=list(mark_data['Regno'])
                         in_data=list(mark_data['Internal'])
                         assgn_data=list(mark_data['Assignment'])
                         qz_data=list(mark_data['Quiz'])
                         lab_data=list(mark_data['Lab'])

                         
                         
                         predicting_mark_window=Toplevel(stf_leaderboard_window)
                         predicting_mark_window.geometry("1200x600")
                         predicting_mark_window.iconbitmap('Datas/src_images/face5.ico')
                         predicting_mark_window.title('Result prediction')
                         predicting_mark_window.resizable(False, False)
                         predicting_mark_window.configure(background='gray')
                         frame = Frame(predicting_mark_window)
                         frame.pack(fill=X,side=TOP)
                         lbt1=Label(frame,text="Result Prediction",font=("Imprint MT Shadow",30))
                         lbt1.pack(fill = X, expand = True,side=TOP)
                         frame1 = Frame(predicting_mark_window)
                         frame1.pack(fill=X,side=TOP)
                         lbl1=Label(frame1,text="Regno",font=('calibri', 26, ' bold '))
                         lbl1.pack(fill = X, expand = True,side=LEFT)
                         lbl2=Label(frame1,text="Internal",font=('calibri', 26, ' bold '))
                         lbl2.pack(fill = X, expand = True,side=LEFT)
                         lbl3=Label(frame1,text="Assignment",font=('calibri', 26, ' bold '))
                         lbl3.pack(fill = X, expand = True,side=LEFT)
                         lbl4=Label(frame1,text="Quiz",font=('calibri', 26, ' bold '))
                         lbl4.pack(fill = X, expand = True,side=LEFT)
                         lbl5=Label(frame1,text="Lab",font=('calibri', 26, ' bold '))
                         lbl5.pack(fill = X, expand = True,side=LEFT)
                         lbl6=Label(frame1,text="Predicted \nMark",font=('calibri', 26, ' bold '))
                         lbl6.pack(fill = X, expand = True,side=LEFT)
                         lbl7=Label(frame1,text="Predicted \nResult",font=('calibri', 26, ' bold '))
                         lbl7.pack(fill = X, expand = True,side=LEFT)
                         def inter_graph():
                              mark_data['Internal'].plot.hist(rwidth=0.8)
                              plt.title("Student Internal Score")
                              plt.xlabel("Internal")
                              plt.ylabel("no of students")
                              plt.rcParams['figure.figsize'] = [10,4]
                              plt.show()
                         def assign_graph():
                              mark_data['Assignment'].plot.hist(rwidth=0.8)
                              plt.title("Student Assignment Score")
                              plt.xlabel("Assignment")
                              plt.ylabel("no of students")
                              plt.rcParams['figure.figsize'] = [10,4]
                              plt.show()
                         def qz_graph():
                              mark_data['Quiz'].plot.hist(rwidth=0.8)
                              plt.title("Student Quiz Score")
                              plt.xlabel("Quiz")
                              plt.ylabel("no of students")
                              plt.rcParams['figure.figsize'] = [10,4]
                              plt.show()
                         def lab_graph():
                              mark_data['Lab'].plot.hist(rwidth=0.8)
                              plt.title("Student Lab Score")
                              plt.xlabel("Lab")
                              plt.ylabel("no of students")
                              plt.rcParams['figure.figsize'] = [10,4]
                              plt.show()     
                         
                         
                         
                         scroll_bar = Scrollbar(predicting_mark_window)     

                         mylist = Listbox(predicting_mark_window,font = ('calibri', 26, ' bold '),background='gray',yscrollcommand = scroll_bar.set)

                         
                         model=pickle.load(open("Datas/mode_file/model.pkl","rb"))
                         #print(name_data,in_data,assgn_data,qz_data,lab_data)
                         pred_list=[]
                         res_list=[]
                         for i in range(len(regno_data)):
                              pred=model.predict([[in_data[i],assgn_data[i],qz_data[i],lab_data[i]]])
                              pred_list.append(int(pred[0]))
                              
                              if pred[0]>=40:
                                   result='PASS'
                              else:
                                   result='FAIL'
                              res_list.append(result)
                              mylist.insert(END,str(regno_data[i])+'              '+str(in_data[i])+'                      '+str(assgn_data[i])+'              '+str(qz_data[i])+'              '+str(lab_data[i])+'              '+str(int(pred[0]))+'              '+str(result))
                              mylist.insert(END,'\n')
                         out_dict1={"prediction":pred_list,"result":res_list}
                         out_df1=pd.DataFrame(out_dict1)

                         def pred_mark_graph():
                              out_df1['prediction'].plot.hist(rwidth=0.8)
                              plt.title("Student predicted Score")
                              plt.xlabel("predicted Score")
                              plt.ylabel("no of students")
                              plt.rcParams['figure.figsize'] = [10,4]
                              plt.show()
                         def pred_result_graph():
                              
                              num = pd.value_counts(out_df1['result'],sort=True).sort_index()
                              num.plot(kind='bar')
                              plt.title("Student predicted result ")
                              plt.xlabel("predicted result")
                              plt.ylabel("no of students")
                              plt.rcParams['figure.figsize'] = [10,4]
                              plt.show()

                         frame2 = Frame(predicting_mark_window)
                         frame2.pack(fill=X,side=TOP)
                         
                         
                         blbl1=Label(frame2,text="To View graph",font=('calibri', 20, ' bold '),fg="black"  ,bg="lightblue", activebackground = "#002266",activeforeground="#80b3ff" )
                         blbl1.pack(fill = X, expand = True,side=LEFT,padx=10)
                         blbl2=Button(frame2,text="show graph",command=inter_graph,font=('calibri', 20, ' bold '),fg="black"  ,bg="lightblue", activebackground = "#002266",activeforeground="#80b3ff")
                         blbl2.pack(fill = X, expand = True,side=LEFT,padx=10)
                         blbl3=Button(frame2,text="show graph",command=assign_graph,font=('calibri', 20, ' bold '),fg="black"  ,bg="lightblue", activebackground = "#002266",activeforeground="#80b3ff")
                         blbl3.pack(fill = X, expand = True,side=LEFT,padx=10)
                         blbl4=Button(frame2,text="show graph",command=qz_graph,font=('calibri', 20, ' bold '),fg="black"  ,bg="lightblue", activebackground = "#002266",activeforeground="#80b3ff")
                         blbl4.pack(fill = X, expand = True,side=LEFT,padx=10)
                         blbl5=Button(frame2,text="show graph",command=lab_graph,font=('calibri', 20, ' bold '),fg="black"  ,bg="lightblue", activebackground = "#002266",activeforeground="#80b3ff")
                         blbl5.pack(fill = X, expand = True,side=LEFT,padx=10)
                         blbl6=Button(frame2,text="show graph",command=pred_mark_graph,font=('calibri', 20, ' bold '),fg="black"  ,bg="lightblue", activebackground = "#002266",activeforeground="#80b3ff")
                         blbl6.pack(fill = X, expand = True,side=LEFT,padx=10)
                         blbl7=Button(frame2,text="show graph",command=pred_result_graph,font=('calibri', 20, ' bold '),fg="black"  ,bg="lightblue", activebackground = "#002266",activeforeground="#80b3ff")
                         blbl7.pack(fill = X, expand = True,side=LEFT,padx=10)

                         scroll_bar.pack( side = RIGHT,fill = Y )
                         mylist.pack( side = TOP, fill = BOTH ,expand=True )
                         scroll_bar.config( command = mylist.yview )
                         
                         predicting_mark_window.mainloop()

                    def openfilename_mk():
                         filename = filedialog.askopenfilename(parent=stf_leaderboard_window,title ="Select csv file")
                         
                         if filename:
                              upvideofile=filename.split('/')
                              if upvideofile[-1] in mark_files:
                                   ms.showerror('Error!','file Already exists',parent=stf_leaderboard_window)
                                   
                              elif filename.endswith('.csv'):
                                   shutil.copy(filename, 'Datas/mark_data')
                                   cr_mark_files=os.listdir('Datas/mark_data')
                                   
                                   
                                   mylist.insert(END,str(len(cr_mark_files))+'. '+str(upvideofile[-1]))
                                   mylist.insert(END,'\n')
                                   ms.showinfo('Success!',upvideofile[-1]+' Upload Successfully!',parent=stf_leaderboard_window)

                                   csv_stu_data=pd.read_csv(filename)
                                   csv_stu_data_col=list(csv_stu_data.columns)

                                                 

                                   if 'Regno' in csv_stu_data_col and 'Internal' in csv_stu_data_col and 'Assignment' in csv_stu_data_col and 'Quiz' in csv_stu_data_col and 'Lab' in csv_stu_data_col:
                                        predicting_mark(filename)
                                   else:
                                        ms.showerror('Error!','Selected file does not have required colums \n Regno, Internal, Assignment,Quiz, Lab .',parent=stf_leaderboard_window)
                                        
                                   
                              else:
                                   ms.showerror('Error!','Selected file is not video file \nMake sure it is csv file.',parent=stf_leaderboard_window)



                    mylist.pack( side = TOP, fill = BOTH ,expand=True )
                    scroll_bar.config( command = mylist.yview )

                    mk_btn1=Button(stf_leaderboard_window,image=import_csv_photo,command=openfilename_mk)
                    mk_btn1.place(x=480,y=300)

                    stf_leaderboard_window.mainloop()


                def staff_assign_down_diff_wn():
                    staff_assign_down_diff_window = Toplevel(stf1_window)
                    staff_assign_down_diff_window.title("Students Assignments Download")

                    staff_assign_down_diff_window.geometry("800x300")
                    staff_assign_down_diff_window.iconbitmap('Datas/src_images/face5.ico')
                    staff_assign_down_diff_window.resizable(False, False)
                    staff_assign_down_diff_window.configure(background='gray')

                    staff_assign_down_diff_window.grid_rowconfigure(0, weight=1)
                    staff_assign_down_diff_window.grid_columnconfigure(0, weight=1)
                    load = Image.open('Datas/src_images/face6.jpg')
                    render = ImageTk.PhotoImage(load)
                    img = Label(staff_assign_down_diff_window, image=render)
                    img.place(x=0, y=0, relwidth=1, relheight=1)
                    lbt1=Label(staff_assign_down_diff_window,text="Staff Assignment Download",font = ('calibri', 24, ' bold '))
                    lbt1.pack(side=TOP)

                    sub_clicked = StringVar()
                    subjectwise_passw = StringVar()
                    pdf_upload_photo = PhotoImage(file = r"Datas/src_images/upload_pdf_logo.png") 
                    Download_logo_photo = PhotoImage(file = r"Datas/src_images/download-logo-12.png")

                    sub_pass_to_download={'Java':'javaisgood',
                                          "Web Technology":'webisgood',
                                          "Software testing":'softwareisgood',
                                          "Mobile app development":'appisgood',
                                          "Fundamentals of datascience":'datascienceisgood',
                                          "Data Structures":'structuresisgood'}

                    assgn_sub_shed_names=['Java',
                                          "Web Technology","Software testing",
                                          "Mobile app development",
                                          "Fundamentals of datascience",
                                          "Data Structures"]

                    pdf_upload_filedict={'Java':'Datas/Students_assign_diff_sub/Java',
                                                 "Web Technology":'Datas/Students_assign_diff_sub/Web Technology',
                                                 "Software testing":'Datas/Students_assign_diff_sub/Software testing',
                                                 "Mobile app development":'Datas/Students_assign_diff_sub/Mobile app development',
                                                 "Fundamentals of datascience":'Datas/Students_assign_diff_sub/Fundamentals of datascience',
                                                 "Data Structures":'Datas/Students_assign_diff_sub/Data Structures'}
                    def multi_asgn_subwise_fol():
                         filename = filedialog.askdirectory(parent=staff_assign_down_diff_window,title ="Select folder")
                         if filename:
                              crt_time=datetime.datetime.now()
                              cre_fol=filename+'/'+sub_clicked.get()+' Assignments-'+crt_time.strftime("%d-%m-%Y")
                              if os.path.exists(cre_fol):
                                   pass
                              else:
                                   os.mkdir(cre_fol)
                              
                              sub_file_path=pdf_upload_filedict[sub_clicked.get()]
                              to_rm_res=ms.askquestion("Delete", "Do you want to clear \nthe datas in the Subject folder", icon='question',parent=staff_assign_down_diff_window)
                              for sub_file_x in os.listdir(sub_file_path):
                                   src_diff_as_fi=os.path.join(sub_file_path,sub_file_x)
                                   shutil.copy(src_diff_as_fi,cre_fol)
                                   if to_rm_res=='yes':
                                        os.remove(src_diff_as_fi)
                              ms.showinfo('Success!','File Downloaded',parent=staff_assign_down_diff_window)
                              
                              
                         
                    def multi_asgn_subwise_download():
                         if subjectwise_passw.get()==sub_pass_to_download[sub_clicked.get()] :
                              multi_asgn_subwise_fol()
                         else:
                              ms.showerror('Oops!','Incorrect password.',parent=staff_assign_down_diff_window)
                         
                    staffup_lbt2=Label(staff_assign_down_diff_window,text="subjectwise password: ",font = ('calibri', 20, ' bold '))
                    staffup_lbt2.place(x=10,y=100)
                    staffup_ent1=Entry(staff_assign_down_diff_window,textvariable=subjectwise_passw,show='*',font = ('calibri', 20, ' bold '),width=35)
                    staffup_ent1.place(x=290,y=100)
                    staffup_lbt3=Label(staff_assign_down_diff_window,text="Select Subject: ",font = ('calibri', 20, ' bold '))
                    staffup_lbt3.place(x=100,y=150)

                    sub_clicked.set( assgn_sub_shed_names[-1] )
                    drop_assgn = OptionMenu( staff_assign_down_diff_window , sub_clicked , *assgn_sub_shed_names )
                    drop_assgn.config( font=('calibri', 18, ' bold '))
                    drop_assgn.place(x=290,y=148)


                    staffup_lbt4=Label(staff_assign_down_diff_window,text="To download the assignments as batch: ",font = ('calibri', 20, ' bold '))
                    staffup_lbt4.place(x=230,y=243)
                    staffup_btn2=Button(staff_assign_down_diff_window,command=multi_asgn_subwise_download,image=Download_logo_photo)
                    staffup_btn2.place(x=700,y=240)


                    staff_assign_down_diff_window.mainloop()



                stf1_btn1 = Button(stf1_window,command=stf_show_att, text="Attendance Report"  ,fg="black"  ,bg="lightblue"  ,width=15  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stf1_btn1.place(x=20, y=100)

                stf1_btn2 = Button(stf1_window,command=stf_record, text="Recording"  ,fg="black"  ,bg="lightblue"  ,width=15  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stf1_btn2.place(x=20, y=200)

                stf1_btn3 = Button(stf1_window,command=stf_class, text="Class"  ,fg="black"  ,bg="lightblue"  ,width=15  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stf1_btn3.place(x=20, y=300)

                stf1_btn4 = Button(stf1_window,command=stf_assignment, text="Assignment"  ,fg="black"  ,bg="lightblue"  ,width=15  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stf1_btn4.place(x=20, y=400)

                stf1_btn5 = Button(stf1_window,command=stf_study_material, text="Study Material"  ,fg="black"  ,bg="lightblue"  ,width=15  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stf1_btn5.place(x=20, y=500)

                stf1_btn6 = Button(stf1_window,command=test, text="Add Student Face"  ,fg="black"  ,bg="lightblue"  ,width=15  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stf1_btn6.place(x=20, y=600)

                stf1_btn7 = Button(stf1_window,command=stf_leaderboard, text="Result Prediction \nand Analysis"  ,fg="black"  ,bg="lightblue"  ,width=15  ,height=2, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stf1_btn7.place(x=510, y=250)

                stf1_btn8 = Button(stf1_window,command=staff_assign_down_diff_wn, text="Student's Assignment"  ,fg="black"  ,bg="lightblue"  ,width=20  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stf1_btn8.place(x=900, y=100)
               
                
                stf1_window.mainloop()
            
            def student_window():
                stud1_window=Toplevel(root)
                stud1_window.title("Student Home Page")                                     
                stud1_window.geometry('1280x700')
                stud1_window.iconbitmap('Datas/src_images/face5.ico')
                stud1_window.state('zoomed')  
                #stud1_window.resizable(False, False)
                stud1_window.configure(background='gray')
                stud1_window.grid_rowconfigure(0, weight=1)
                stud1_window.grid_columnconfigure(0, weight=1)
                load = Image.open('Datas/src_images/face_reco.jpg')
                render = ImageTk.PhotoImage(load)
                img = Label(stud1_window, image=render)
                img.place(x=0, y=0, relwidth=1, relheight=1)

                video_upload_photo = ImageTk.PhotoImage(file = "Datas/src_images/upload_btn_logo1.png")

                pdf_upload_photo = ImageTk.PhotoImage(file = "Datas/src_images/upload_pdf_logo.png")

                stf_study_upload_photo = ImageTk.PhotoImage(file = "Datas/src_images/upload_pdf_logo_stm.png")
                def TrackImages():
                    recognizer = cv2.face.LBPHFaceRecognizer_create()
                    recognizer.read("Atten_datas/TrainingImageLabel/Trainner.yml")
                    harcascadePath="Datas/haarcascade_frontalface_default.xml"
                    faceCascade = cv2.CascadeClassifier(harcascadePath);    
                    df=pd.read_csv('Atten_datas/AuthorisedDetails/AuthorisedDetails.csv')
                    cam = cv2.VideoCapture(0)
                    font = cv2.FONT_HERSHEY_SIMPLEX        
                    col_names =  ['Id','Name','Date','Time']
                    attendance = pd.DataFrame(columns = col_names)    
                    while True:
                        ret, im =cam.read()
                        gray=cv2.cvtColor(im,cv2.COLOR_BGR2GRAY)
                        faces=faceCascade.detectMultiScale(gray, 1.2,5)    
                        for(x,y,w,h) in faces:
                            cv2.rectangle(im,(x,y),(x+w,y+h),(225,0,0),2)
                            Id, conf = recognizer.predict(gray[y:y+h,x:x+w])                                   
                            if(conf < 50):
                                ts = time.time()      
                                date = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
                                calibritamp = datetime.datetime.fromtimestamp(ts).strftime('%H:%M:%S')
                                aa=df.loc[df['Id'] == Id]['Name'].values
                                tt=str(Id)+"-"+aa[0]
                                attendance.loc[len(attendance)] = [Id,aa[0],date,calibritamp]
                                
                            else:
                                Id='Unknown'                
                                tt=str(Id)  
                            if(conf > 75):
                                Id='Unknown'                
                                tt=str(Id)
                                
                           
                            cv2.putText(im,str(tt),(x,y+h), font, 1,(255,255,255),2)        
                        attendance=attendance.drop_duplicates(subset=['Id'],keep='first')    
                        cv2.imshow('image Detected: Enter q to quit',im) 
                        if (cv2.waitKey(1)==ord('q')):
                            break
                    ts = time.time()      
                    date = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
                    calibritamp = datetime.datetime.fromtimestamp(ts).strftime('%H:%M:%S')
                    Hour,Minute,Second=calibritamp.split(":")
                    fileName="Atten_datas/Attendance_sheet/"+"Attendance_"+date+"_"+".csv"
                    attendance.to_csv(fileName,index=False)
                    cam.release()
                    cv2.destroyAllWindows()
                    #print(attendance)


                def stud_asgn_upload():
                    stud_assignments_up_window = Toplevel(stud1_window)
                    stud_assignments_up_window.title("Students Assignments Upload")

                    stud_assignments_up_window.geometry("1000x440")
                    stud_assignments_up_window.iconbitmap('Datas/src_images/face5.ico')
                    stud_assignments_up_window.resizable(False, False)
                    stud_assignments_up_window.configure(background='gray')

                    sub_clicked = StringVar()

                    pdf_upload_photo = PhotoImage(file = r"Datas/src_images/upload_pdf_logo.png") 
                    scroll_bar = Scrollbar(stud_assignments_up_window)

                    scroll_bar.pack( side = RIGHT,fill = Y )

                    mylist = Listbox(stud_assignments_up_window,font = ('calibri', 22, ' bold '),background='gray',yscrollcommand = scroll_bar.set )
                    assgn_sub_shed_names=['Java',"Web Technology","Software testing","Mobile app development","Fundamentals of datascience","Data Structures"]
                    pdf_upload_filedict={'Java':'Datas/Students_assign_diff_sub/Java',
                                                 "Web Technology":'Datas/Students_assign_diff_sub/Web Technology',
                                                 "Software testing":'Datas/Students_assign_diff_sub/Software testing',
                                                 "Mobile app development":'Datas/Students_assign_diff_sub/Mobile app development',
                                                 "Fundamentals of datascience":'Datas/Students_assign_diff_sub/Fundamentals of datascience',
                                                 "Data Structures":'Datas/Students_assign_diff_sub/Data Structures'}

                    Assignments_pdf_files=[]

                    for upload_sub_key,upload_sub_path in pdf_upload_filedict.items():
                            upload_sub_val_files=os.listdir(upload_sub_path)
                            for up_sub_val_file in upload_sub_val_files:
                                    Assignments_pdf_files.append(up_sub_val_file+'  ('+upload_sub_key+')')

                    for linenum,line in enumerate(Assignments_pdf_files):
                        mylist.insert(END,str(linenum+1)+'. '+str(line))
                        mylist.insert(END,'\n')

                    def openfile_Students_up_Assignments_pdf(pdf_upload_filename):
                            filename = filedialog.askopenfilename(parent=stud_assignments_up_window,title ="Select pdf")
                            if filename:
                                    uppdffile=filename.split('/')
                                    cr_Assignments_pdf_files=os.listdir(pdf_upload_filename)
                                    if uppdffile[-1] in cr_Assignments_pdf_files:
                                            ms.showerror('Error!','file Already exists',parent=stud_assignments_up_window)
                                    elif filename.endswith('.pdf'):
                                            shutil.copy(filename, pdf_upload_filename)
                                            ms.showinfo('Success!',uppdffile[-1]+' Upload Successfully!',parent=stud_assignments_up_window)
                                            mylist.insert(END,str(len(Assignments_pdf_files)+1)+'. '+str(uppdffile[-1])+' (uploaded to --> '+sub_clicked.get()+' )')
                                            mylist.insert(END,'\n')
                                    else:
                                            ms.showerror('Error!','Selected file is not pdf file \nMake sure it is pdf.',parent=stud_assignments_up_window)

                    def openfile_Students_up_diff_Assignments_pdf():
                            

                            pdf_up_filename_val=pdf_upload_filedict[sub_clicked.get()]
                            
                            openfile_Students_up_Assignments_pdf(pdf_up_filename_val)
                            

                    mylist.pack( side = TOP, fill = X ,expand=True )
                    scroll_bar.config( command = mylist.yview )



                    frame = Frame(stud_assignments_up_window,bg='grey')
                    frame.pack(fill=X,side=BOTTOM)

                    stasgnup_btn1=Label(frame,text="<--UPLOAD",font=('calibri', 22, ' bold '),fg="black"  ,bg="grey"  ,width=10  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" )
                    stasgnup_btn1.pack(side=RIGHT,padx=15)

                    stasgnup_btn2=Button(frame,image=pdf_upload_photo,command=openfile_Students_up_diff_Assignments_pdf)
                    stasgnup_btn2.pack(fill = X,side=RIGHT)



                    lbt1=Label(frame,text="<--Select Subject ",font = ('calibri', 22, ' bold '),bg='grey')
                    lbt1.pack(fill = X,side=RIGHT)


                    sub_clicked.set( assgn_sub_shed_names[-1] )
                    drop_assgn = OptionMenu( frame , sub_clicked , *assgn_sub_shed_names )
                    drop_assgn.config( font=('calibri', 20, ' bold '))
                    drop_assgn.pack(fill = X,side=RIGHT)


                    stud_assignments_up_window.mainloop()

                  
                def class_recoding_download():
                  stud_videos_down_window = Toplevel(stud1_window)
                  stud_videos_down_window.title("Students Class recording download")

                  stud_videos_down_window.geometry("600x450")
                  stud_videos_down_window.iconbitmap('Datas/src_images/face5.ico')
                  stud_videos_down_window.resizable(False, False)
                  stud_videos_down_window.configure(background='gray')
                  Download_logo_photo = PhotoImage(file = r"Datas/src_images/download-logo-12.png")

                  scroll_bar = Scrollbar(stud_videos_down_window)
                  scroll_bar.pack( side = RIGHT,fill = Y )
                  mylist = Listbox(stud_videos_down_window,font = ('calibri', 22, ' bold '),background='gray',yscrollcommand = scroll_bar.set )



                  videos_down_files=os.listdir('Datas/videos')
                  for linenum,line in enumerate(videos_down_files):
                          mylist.insert(END,str(linenum+1)+'. '+str(line))
                          mylist.insert(END,'\n')
                  mylist.pack( side = TOP, fill = BOTH ,expand=True )
                  scroll_bar.config( command = mylist.yview )
                  frame = Frame(stud_videos_down_window)
                  frame.pack(fill=X,side=BOTTOM)
                  lbt1=Label(frame,text="Click here-->",font = ('calibri', 22, ' bold '))
                  lbt1.pack(fill = X,side=LEFT)


                  def download_recordings_file():
                       filename = filedialog.asksaveasfilename(parent=stud_videos_down_window,title ="Select Image")
                       if filename:
                            if filename.endswith('.mp4')or filename.endswith('.mkv'):
                                shutil.copy('Datas/videos/'+clicked.get(),filename)
                                ms.showinfo('Success!','File Downloaded',parent=stud_videos_down_window)
                            elif filename:
                                 fullfilename=filename+'.mp4'
                                 shutil.copy('Datas/videos/'+clicked.get(),fullfilename)
                                 ms.showinfo('Success!','File Downloaded',parent=stud_videos_down_window)
                            else:
                                 ms.showerror('Error!','Selected file is not video file \nMake sure it is mp4 or mkv file.',parent=stud_videos_down_window)

                  stasgnup_btn1=Button(frame,image=Download_logo_photo,command=download_recordings_file)
                  stasgnup_btn1.pack(side=LEFT)
                  clicked = StringVar()
                  clicked.set( videos_down_files[-1] )
                  drop_videos = OptionMenu( frame , clicked , *videos_down_files )
                  drop_videos.config( font=('calibri', 20, ' bold '))
                  drop_videos.pack(fill = X,side=LEFT)
                  stud_videos_down_window.mainloop()


                def class_assignments_download():
                    stud_assgn_down_window = Toplevel(stud1_window)
                    stud_assgn_down_window.title("Students Assignments Download")

                    stud_assgn_down_window.geometry("600x450")
                    stud_assgn_down_window.iconbitmap('Datas/src_images/face5.ico')
                    stud_assgn_down_window.resizable(False, False)
                    stud_assgn_down_window.configure(background='gray')
                    Download_logo_photo = PhotoImage(file = r"Datas/src_images/download-logo-12.png")

                    scroll_bar = Scrollbar(stud_assgn_down_window)
                    scroll_bar.pack( side = RIGHT,fill = Y )
                    mylist = Listbox(stud_assgn_down_window,font = ('calibri', 22, ' bold '),background='gray',yscrollcommand = scroll_bar.set )



                    assgn_down_files=os.listdir('Datas/Assignments_Upload')
                    for linenum,line in enumerate(assgn_down_files):
                        mylist.insert(END,str(linenum+1)+'. '+str(line))
                        mylist.insert(END,'\n')
                    mylist.pack( side = TOP, fill = BOTH ,expand=True )
                    scroll_bar.config( command = mylist.yview )
                    frame = Frame(stud_assgn_down_window)
                    frame.pack(fill=X,side=BOTTOM)
                    lbt1=Label(frame,text="Click here-->",font = ('calibri', 22, ' bold '))
                    lbt1.pack(fill = X,side=LEFT)


                    def download_Assignments_file():
                         filename = filedialog.asksaveasfilename(parent=stud_assgn_down_window,title ="Select Download path")
                         if filename:
                              if filename.endswith('.pdf'):
                                  shutil.copy('Datas/Assignments_Upload/'+clicked.get(),filename)
                                  ms.showinfo('Success!','File Downloaded',parent=stud_assgn_down_window)
                              elif filename:
                                   fullfilename=filename+'.pdf'
                                   shutil.copy('Datas/Assignments_Upload/'+clicked.get(),fullfilename)
                                   ms.showinfo('Success!','File Downloaded',parent=stud_assgn_down_window)
                              else:
                                   ms.showerror('Error!','Selected file is not video file \nMake sure it is mp4 or mkv file.',parent=stud_assgn_down_window)

                    stasgnup_btn1=Button(frame,image=Download_logo_photo,command=download_Assignments_file)
                    stasgnup_btn1.pack(side=LEFT)
                    clicked = StringVar()
                    clicked.set( assgn_down_files[-1] )
                    drop_assgn = OptionMenu( frame , clicked , *assgn_down_files )
                    drop_assgn.config( font=('calibri', 20, ' bold '))
                    drop_assgn.pack(fill = X,side=LEFT)
                    stud_assgn_down_window.mainloop()

                def zoom_class_meeting_widget_window():
                    stud_zoom_class_window=Toplevel(stud1_window)
                    stud_zoom_class_window.title("Online class")

                                                 
                    stud_zoom_class_window.geometry('1280x700')
                    stud_zoom_class_window.iconbitmap('Datas/src_images/face5.ico')
                    stud_zoom_class_window.resizable(False, False)
                    stud_zoom_class_window.configure(background='gray')
                    stud_zoom_class_window.grid_rowconfigure(0, weight=1)
                    stud_zoom_class_window.grid_columnconfigure(0, weight=1)
                    load = Image.open('Datas/src_images/face6.jpg')
                    render = ImageTk.PhotoImage(load)
                    img = Label(stud_zoom_class_window, image=render)
                    img.place(x=0, y=0, relwidth=1, relheight=1)
                    stud_zoom_label = Label(stud_zoom_class_window, text="Choose the subject" ,bg="lightblue"  ,fg="black"  ,width=45  ,height=1,font=('calibri', 40, 'bold')) 
                    stud_zoom_label.pack(side=TOP)

                    def zoom_class_deploy(zoom_list_file):
                        try:
                            import  datetime, time, subprocess, csv, os, webbrowser

                            try:
                                import pyautogui
                                import openpyxl
                                from PIL import Image
                            except ModuleNotFoundError as err:
                                print("not installed modules please go through the read me files, press anything to exit")
                                input()
                            #enabling mouse fail safe

                              
                            pyautogui.FAILSAFE = True

                            #copying data from excel sheet to the program
                            meetings = []
                            wb = openpyxl.load_workbook(zoom_list_file)
                            sheet = wb['Sheet1']

                            for i in sheet.iter_rows(values_only = True):
                                if i[0] != None:   
                                    meetings.append(i)
                            meetings.pop(0)
                            meetings.sort()


                            #function to manualy join if no link is provided
                            def manualjoin(id, password = ""):
                                  
                                time.sleep(3)

                                #locating the zoom app
                                while True:
                                    var = pyautogui.locateOnScreen('Datas/zoom_meeting_items/zoom_src_images/final.png', confidence=0.9)
                                    if var != None:
                                        pyautogui.click(var)
                                        break
                                    elif (time.time() - cur) >= 120:
                                        print("App Not opened")
                                        break
                                    #check every 30 secs
                                    time.sleep(30)

                                time.sleep(3)

                                #entering the meeting id
                                pyautogui.typewrite(id)

                                #disabling video source
                                var = pyautogui.locateOnScreen('Datas/zoom_meeting_items/zoom_src_images/videooff.png', confidence=0.9)
                                pyautogui.click(var)

                                #clicking the join button
                                var = pyautogui.locateOnScreen('Datas/zoom_meeting_items/zoom_src_images/join.png', confidence=0.9)
                                join = (var[0] + 75, var[1] + 10, var[2], var[3])
                                pyautogui.moveTo(pyautogui.center(join))
                                pyautogui.click(join)

                                time.sleep(3)

                                #checking and entering if meeting password is enabled
                                if pyautogui.locateOnScreen('Datas/zoom_meeting_items/zoom_src_images/password.png', confidence=0.9) != None :
                                    pyautogui.typewrite(password)
                                    var = pyautogui.locateOnScreen('Datas/zoom_meeting_items/zoom_src_images/joinmeeting.png', confidence=0.9)
                                    pyautogui.click(var)

                                return



                            def linkjoin(link):
                                #open the given link in web browser
                                webbrowser.open(link)
                                start = time.time()
                                time.sleep(3)
                                while True:
                                    var = pyautogui.locateOnScreen('Datas/zoom_meeting_items/zoom_src_images/openlink.png', confidence=0.9)
                                    if var != None:
                                        pyautogui.click(var)
                                        break
                                    var = pyautogui.locateOnScreen('Datas/zoom_meeting_items/zoom_src_images/openzoom.png', confidence=0.9)
                                    if var != None:
                                        pyautogui.click(var)
                                        break
                                    elif (time.time() - start) >= 120:
                                        print("link " + link + " not opened")
                                        break
                                    time.sleep(3)
                                return


                            #Iterating through the meeting list to jointate the specified time
                            for i in range(len(meetings)):
                                curmeeting = meetings[i]

                                #Setting the meeting Times
                                cur = round(time.time(), 0)
                                temp = curmeeting[0].timestamp()

                                #join a minute early for later scheduled class
                                if(cur <= temp ):
                                    print("class will be active for ", end ="")
                                    print(datetime.timedelta(seconds = (temp - cur) - 60))
                                    #time.sleep(temp - cur - 60)
                                #if more than 5 minutes have passed already
                                elif (cur > temp):
                                    print("skipped meeting " + str(i + 1))
                                    continue
                                      
                                #var = os.system("taskkill /f /im Zoom.exe")
                                  
                                  
                                #check if link is provided
                                if curmeeting[1] != None:
                                    linkjoin(str(curmeeting[1]))
                                #check if 
                                if curmeeting[2] != None:
                                    subfolders = [ f.path for f in os.scandir("C:\\Users") if f.is_dir() ]
                                    #opening the zoom app, if you are running on a different OS or the path is different
                                    #change the path here 
                                    for i in subfolders:
                                        if os.path.isfile(i + "\\AppData\\Roaming\\Zoom\\bin\\Zoom.exe"):
                                            subprocess.Popen(i + "\\AppData\\Roaming\\Zoom\\bin\\Zoom.exe")  
                                    manualjoin(str(curmeeting[2]), str(curmeeting[3]))
                                  
                                else:
                                    print("data insufficient, press anything to exit")
                                    input()
                                    exit()

                                time.sleep(5)
                                #check whether the class has started and enabling audio
                                while True:
                                    if pyautogui.locateOnScreen('Datas/zoom_meeting_items/zoom_src_images/audioenable.png', confidence=0.9) != None :
                                        var = pyautogui.locateOnScreen('Datas/zoom_meeting_items/zoom_src_images/audioenable.png', confidence=0.9)
                                        pyautogui.click(var)
                                        break
                                    elif pyautogui.locateOnScreen('Datas/zoom_meeting_items/zoom_src_images/leave.png', confidence=0.9) != None :
                                        var = pyautogui.locateOnScreen('Datas/zoom_meeting_items/zoom_src_images/leave.png', confidence=0.9)
                                        pyautogui.click(var)
                                        break
                                    elif (time.time() - cur) >= 30 * 60:
                                        os.system("taskkill /f /im Zoom.exe")
                                        break
                                    time.sleep(5)

                                #check whether the mic is muted, if not muted
                                pyautogui.moveTo(x = 900, y = 900, duration = 0.25)
                                if pyautogui.locateOnScreen('Datas/zoom_meeting_items/zoom_src_images/mute.png', confidence=0.9) != None :
                                    var = pyautogui.locateOnScreen('Datas/zoom_meeting_items/zoom_src_images/mute.png', confidence=0.9)
                                    pyautogui.click(var)


                            #program has finished all classes and exits
                              
                            ms.showinfo('Sorry','Meeting has ended',parent=stud_zoom_class_window)
                            var = os.system("taskkill /f /im Zoom.exe")

                              
                              
                              
                        except:
                            ms.showerror('Error!','Something went Wrong Try Again.',parent=stud_zoom_class_window)
                            stud_zoom_class_window.destroy()

                         
                    def Java_func():
                        zoom_class_deploy('Datas/zoom_meeting_items/zoom_meet_xlfiles/Java_meet.xlsx')
                         

                    def Web_Technology_func():
                        zoom_class_deploy('Datas/zoom_meeting_items/zoom_meet_xlfiles/Web_Technology_meet.xlsx')

                    def Software_testing_func():
                         
                        zoom_class_deploy('Datas/zoom_meeting_items/zoom_meet_xlfiles/Software testing_meet.xlsx')

                    def Mobile_app_development_func():
                        zoom_class_deploy('Datas/zoom_meeting_items/zoom_meet_xlfiles/Mobile_app_development_meet.xlsx')

                    def Fundamentals_of_datascience_func():
                        zoom_class_deploy('Datas/zoom_meeting_items/zoom_meet_xlfiles/Fundamentals_of_datascience_meet.xlsx')

                    def Data_Structures_func():
                        zoom_class_deploy('Datas/zoom_meeting_items/zoom_meet_xlfiles/Data_Structures_meet.xlsx')
                         
                    stud_sub_zoom_btn1 = Button(stud_zoom_class_window,command=Java_func, text="Java"  ,fg="black"  ,bg="lightblue"  ,width=15  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                    stud_sub_zoom_btn1.place(x=20, y=100)

                    stud_sub_zoom_btn2 = Button(stud_zoom_class_window,command=Web_Technology_func, text="Web Technology"  ,fg="black"  ,bg="lightblue"  ,width=15  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                    stud_sub_zoom_btn2.place(x=20, y=200)

                    stud_sub_zoom_btn3 = Button(stud_zoom_class_window,command=Software_testing_func, text="Software testing"  ,fg="black"  ,bg="lightblue"  ,width=15  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                    stud_sub_zoom_btn3.place(x=20, y=300)

                    stud_sub_zoom_btn4 = Button(stud_zoom_class_window,command=Mobile_app_development_func, text="Mobile app development"  ,fg="black"  ,bg="lightblue"  ,width=25  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                    stud_sub_zoom_btn4.place(x=20, y=400)

                    stud_sub_zoom_btn5 = Button(stud_zoom_class_window,command=Fundamentals_of_datascience_func, text="Fundamentals of datascience"  ,fg="black"  ,bg="lightblue"  ,width=25  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                    stud_sub_zoom_btn5.place(x=20, y=500)

                    stud_sub_zoom_btn6 = Button(stud_zoom_class_window,command=Data_Structures_func, text="Data Structures"  ,fg="black"  ,bg="lightblue"  ,width=15  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                    stud_sub_zoom_btn6.place(x=20, y=600)               
                                    
                    stud_zoom_class_window.mainloop()

                    var = os.system("taskkill /f /im Zoom.exe")


                def download_study_materials_file():
                    stud_study_mat_down_window = Toplevel(stud1_window)
                    stud_study_mat_down_window.title("Students study materials Download")

                    stud_study_mat_down_window.geometry("600x450")
                    stud_study_mat_down_window.iconbitmap('Datas/src_images/face5.ico')
                    stud_study_mat_down_window.resizable(False, False)
                    stud_study_mat_down_window.configure(background='gray')
                    Download_logo_photo = PhotoImage(file = r"Datas/src_images/download-logo-12.png")

                    scroll_bar = Scrollbar(stud_study_mat_down_window)
                    scroll_bar.pack( side = RIGHT,fill = Y )
                    mylist = Listbox(stud_study_mat_down_window,font = ('calibri', 22, ' bold '),background='gray',yscrollcommand = scroll_bar.set )



                    study_mat_down_files=os.listdir('Datas/study_materials')
                    for linenum,line in enumerate(study_mat_down_files):
                        mylist.insert(END,str(linenum+1)+'. '+str(line))
                        mylist.insert(END,'\n')
                    mylist.pack( side = TOP, fill = BOTH ,expand=True )
                    scroll_bar.config( command = mylist.yview )
                    frame = Frame(stud_study_mat_down_window)
                    frame.pack(fill=X,side=BOTTOM)
                    lbt1=Label(frame,text="Click here-->",font = ('calibri', 22, ' bold '))
                    lbt1.pack(fill = X,side=LEFT)


                    def download_study_materials_file():
                         filename = filedialog.asksaveasfilename(parent=stud_study_mat_down_window,title ="Select Image")
                         if filename:
                              if filename.endswith('.pdf'):
                                  shutil.copy('Datas/study_materials/'+clicked.get(),filename)
                                  ms.showinfo('Success!','File Downloaded',parent=stud_zoom_class_window)
                              elif filename:
                                   fullfilename=filename+'.pdf'
                                   shutil.copy('Datas/study_materials/'+clicked.get(),fullfilename)
                                   ms.showinfo('Success!','File Downloaded',parent=stud_zoom_class_window)
                              else:
                                   ms.showerror('Error!','Selected file is not pdf file \nMake sure it is pdf file.',parent=stud_zoom_class_window)

                    stasgnup_btn1=Button(frame,image=Download_logo_photo,command=download_study_materials_file)
                    stasgnup_btn1.pack(side=LEFT)
                    clicked = StringVar()
                    clicked.set( study_mat_down_files[-1] )
                    drop_study_mat = OptionMenu( frame , clicked , *study_mat_down_files )
                    drop_study_mat.config( font=('calibri', 20, ' bold '))
                    drop_study_mat.pack(fill = X,side=LEFT)
                    stud_study_mat_down_window.mainloop()
                     

                     


                message = Label(stud1_window, text="Welcome Student : "+self.username.get() ,bg="lightblue"  ,fg="black"  ,width=45  ,height=1,font=('calibri', 40, 'bold')) 
                message.pack(side=TOP)
                
                stud1_btn1 = Button(stud1_window,command=zoom_class_meeting_widget_window, text="Join class"  ,fg="black"  ,bg="lightblue"  ,width=20  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stud1_btn1.place(x=20, y=100)

                stud1_btn2 = Button(stud1_window,command=TrackImages, text="Attendance"  ,fg="black"  ,bg="lightblue"  ,width=20  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stud1_btn2.place(x=20, y=200)

                stud1_btn3 = Button(stud1_window,command=class_recoding_download,text="Class Recordings"  ,fg="black"  ,bg="lightblue"  ,width=20  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stud1_btn3.place(x=20, y=300)

                stud1_btn4 = Button(stud1_window,command=class_assignments_download, text="Download Assignment"  ,fg="black"  ,bg="lightblue"  ,width=20  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stud1_btn4.place(x=20, y=400)

                stud1_btn5 = Button(stud1_window,command=stud_asgn_upload, text="Upload Assignment"  ,fg="black"  ,bg="lightblue"  ,width=20  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stud1_btn5.place(x=20, y=500)

                stud1_btn6 = Button(stud1_window,command=download_study_materials_file,text="Study Material"  ,fg="black"  ,bg="lightblue"  ,width=20  ,height=1, activebackground = "#002266",activeforeground="#80b3ff" ,font=('calibri', 24, ' bold '))
                stud1_btn6.place(x=20, y=600)
                        
                        
                stud1_window.mainloop()



            if result[0][0]=='staff':
                staff_window()
            elif result[0][0]=='student':
                student_window()

        
        else:
            ms.showerror('Oops!','Username Not Found.')
            
    def new_user(self):
    	#Establish Connection
        with sqlite3.connect('Datas/Account_database.db') as db:
            c = db.cursor()

        #Find Existing username if any take proper action
        find_user = ('SELECT username FROM user WHERE username = ?')
        c.execute(find_user,[(self.n_username.get())])        
        if c.fetchall():
            ms.showerror('Error!','Username Taken Try a Diffrent One.')
        elif self.n_secret_key.get()!=secret_key:
            ms.showerror('Error!','Secret Key is Wrong Try Again.')

        else:
            
            self.log()
            #Create New Account 
            insert = 'INSERT INTO user(username,password,category) VALUES(?,?,?)'
            c.execute(insert,[(self.n_username.get()),(self.n_password.get()),(self.n_category.get())])
            db.commit()
            ms.showinfo('Success!','Account Created!')

        #Frame Packing Methords
    def log(self):
        self.username.set('')
        self.password.set('')
        self.crf.pack_forget()
        self.head['text'] = 'LOGIN'
        self.logf.pack()
    def cr(self):
        self.n_username.set('')
        self.n_password.set('')
        self.logf.pack_forget()
        self.head['text'] = 'Create Account'
        self.crf.pack()
        
    #Draw Widgets
    def widgets(self):
        self.master.iconbitmap('Datas/src_images/face5.ico')
        self.head = Label(self.master,text = 'LOGIN',font = ('',35),pady = 10,bg="lightblue" )
        self.head.pack()
        self.logf = Frame(self.master,padx =10,pady = 10,bg="lightblue" )
        Label(self.logf,text = 'Username: ',font = ('',20),pady=5,padx=5,bg="lightblue" ).grid(sticky = W)
        Entry(self.logf,textvariable = self.username,bd = 5,font = ('',15),bg="lightblue" ).grid(row=0,column=1)
        Label(self.logf,text = 'Password: ',font = ('',20),pady=5,padx=5,bg="lightblue" ).grid(sticky = W)
        Entry(self.logf,textvariable = self.password,bd = 5,font = ('',15),show = '*',bg="lightblue" ).grid(row=1,column=1)
        Button(self.logf,text = '      Login     ',bd = 3 ,font = ('',15),padx=5,pady=5,command=self.login,bg="lightblue" ).grid(row=2,column=1)
        Button(self.logf,text = ' Create Account ',bd = 3 ,font = ('',15),padx=5,pady=5,command=self.cr,bg="lightblue" ).grid(row=2,column=0)
        self.logf.pack()
        
        self.crf = Frame(self.master,padx =10,pady = 10,bg="lightblue" )
        Label(self.crf,text = 'Username: ',font = ('',20),pady=5,padx=5,bg="lightblue" ).grid(sticky = W)
        Entry(self.crf,textvariable = self.n_username,bd = 5,font = ('',15),bg="lightblue" ).grid(row=0,column=1)
        Label(self.crf,text = 'Password: ',font = ('',20),pady=5,padx=5,bg="lightblue" ).grid(sticky = W)
        Entry(self.crf,textvariable = self.n_password,bd = 5,font = ('',15),show = '*',bg="lightblue" ).grid(row=1,column=1)
        Label(self.crf,text = 'Staff Key: ',font = ('',20),pady=5,padx=5,bg="lightblue" ).grid(sticky = W)
        Entry(self.crf,textvariable = self.n_secret_key,bd = 5,font = ('',15),show = '*',bg="lightblue" ).grid(row=2,column=1)
        Label(self.crf,text = 'Category: ',font = ('',20),pady=5,padx=5,bg="lightblue" ).grid(sticky = W)
        #Radiobutton(self.crf,text="Staff",variable = self.n_category,bd = 5,font = ('',15),bg="lightblue",value="staff").grid(row=3,column=1,padx=(50,50))
        Radiobutton(self.crf,text="Staff",variable = self.n_category,bd = 5,font = ('',15),bg="lightblue",value="staff").place(x=160,y=140)
        Radiobutton(self.crf,text="Student",variable = self.n_category,bd = 5,font = ('',15),bg="lightblue",value="student").place(x=280,y=140)
        Button(self.crf,text = 'Create Account',bd = 3 ,font = ('',15),padx=5,pady=5,bg="lightblue" ,command=self.new_user).grid(row=4,column=1,padx=(70,0))
        Button(self.crf,text = ' Go to Login  ',bd = 3 ,font = ('',15),padx=5,pady=5,bg="lightblue" ,command=self.log).grid(row=4,column=0)
try:    
    if __name__ == '__main__':
        root = Tk()
        root.title('Login Form')
        #root.geometry('400x350+300+300')
        root.geometry('720x360')
        root.resizable(False, False)
        loadmain = Image.open('Datas/src_images/face_login.jpg')
        rendermain = ImageTk.PhotoImage(loadmain)
        imgmain = Label(root, image=rendermain)
        imgmain.place(x=0, y=0, relwidth=1, relheight=1)
        main(root)
        root.mainloop()
except:
    ms.showerror('Error!','Sorry Error Occured \nKindy Restart the App')
    root.destroy()
